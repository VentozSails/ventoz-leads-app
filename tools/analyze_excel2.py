#!/usr/bin/env python3
"""Deeper analysis - find totaal row and problematic rows."""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\irvhu\Desktop\2026 Ventoz VOORRAAD Zeilen.xlsx', data_only=True)
ws = wb.active

# Find all rows where any cell contains 'totaal'
print("=== ROWS WITH 'totaal' ===")
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    for ci, c in enumerate(cells[:30]):
        if c is not None and 'totaal' in str(c).lower():
            vals = []
            for x in cells[:15]:
                vals.append(str(x)[:20] if x is not None else '')
            print(f'Row {ridx}, match in col {ci}: ' + ' | '.join(vals))
            break

# Now let's look at the specific database rows that had shifted columns
# From the CSV export, the problematic items were:
# id 11 (ean_code=53, kleur=lichtblauw in wrong col)
# id 33 (ean_code=72, grijs en wit in wrong col)  
# id 79 (ean_code=45, wit in wrong col)
# id 83 (ean_code=51, luff 518 cm in wrong col)
# id 93 (ean_code=96, wit + gele patches in wrong col)
# id 95 (ean_code=97, wit in wrong col)
#
# These all have artikelnummer values that look like they were put in the ean_code field
# Let's find them in Excel by looking for those artikelnummers

print("\n=== LOOKING FOR SHIFTED ROWS ===")
target_arts = ['53', '72', '45', '51', '96', '97']
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    artikelnr = str(cells[3] or '').strip()
    if artikelnr in target_arts:
        vals = []
        for x in cells[:30]:
            vals.append(str(x)[:25] if x is not None else '')
        print(f'\nRow {ridx} (art={artikelnr}):')
        for ci, v in enumerate(vals):
            if v:
                col_letter = chr(65 + ci) if ci < 26 else chr(64 + ci // 26) + chr(65 + ci % 26)
                print(f'  Col {ci:2d} ({col_letter}): {v}')

# Count actual product rows (non-empty col B)
product_count = 0
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    product = str(cells[1] or '').strip()
    if product and product.lower() not in ('totaal', 'totalen'):
        product_count += 1
print(f'\n\nTotal product rows (non-empty col B, excl totaal): {product_count}')
