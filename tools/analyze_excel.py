#!/usr/bin/env python3
"""Analyze Excel structure to find problematic rows."""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\irvhu\Desktop\2026 Ventoz VOORRAAD Zeilen.xlsx', data_only=True)
ws = wb.active
print(f'Total rows: {ws.max_row}, Total cols: {ws.max_column}')
print()

header_cells = [str(c.value or '') for c in ws[1]]
for i, h in enumerate(header_cells):
    col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
    print(f'  Col {i:2d} ({col_letter}): {h[:60]}')
print()

print("=== FLAGGED ROWS ===")
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)

    product = str(cells[1] or '').strip().lower()
    artikelnr = str(cells[3] or '').strip()
    ean = str(cells[7] or '').strip()

    flags = []
    if product == 'totaal':
        flags.append('TOTAAL-ROW')
    if '#REF' in artikelnr:
        flags.append('REF-IN-ARTNR')
    if '#REF' in ean:
        flags.append('REF-IN-EAN')

    box_names = ['wit kleinste', 'wit middel', 'midden', 'groot', 'lang standaard',
                 'lang groot', 'splash', 'mk ii', 'valk']
    if product in box_names:
        flags.append('BOX-ROW')

    if flags:
        vals = []
        for c in cells[:10]:
            vals.append(str(c)[:25] if c is not None else '')
        print(f'Row {ridx:3d}: ' + ' | '.join(vals))
        print(f'         FLAGS: ' + ', '.join(flags))
        print()

print("\n=== LAST 15 ROWS ===")
all_rows = list(ws.iter_rows(min_row=2, values_only=True))
start = max(0, len(all_rows) - 15)
for idx, row in enumerate(all_rows[start:], start=start + 2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    vals = []
    for c in cells[:10]:
        vals.append(str(c)[:25] if c is not None else '')
    print(f'Row {idx:3d}: ' + ' | '.join(vals))
