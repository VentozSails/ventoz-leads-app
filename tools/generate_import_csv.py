#!/usr/bin/env python3
"""Generate CSV files from Ventoz Excel files for import into the app.

Excel structure (2026 Ventoz VOORRAAD Zeilen.xlsx):
  Col A (0):  categorie (filled sporadically, propagates down)
  Col B (1):  product naam (filled for first variant, empty for subsequent)
  Col C (2):  kleur
  Col D (3):  artikelnummer
  Col E (4):  gewicht zeil (gram)
  Col F (5):  gewicht verpakking (gram)
  Col G (6):  totaal gewicht
  Col H (7):  EAN code
  Col I (8):  opmerking
  Col J (9):  huidige voorraad
  Col K (10): aantal besteld
  Col L (11): minimale voorraad 2025
  Col M (12): minimale voorraad 2026
  Col N (13): inkoop op Wilfer factuur
  Col O (14): vervoer
  Col P (15): extra vliegtuig kosten
  Col Q (16): invoertax + administratie
  Col R (17): (empty)
  Col S (18): INKOOP PRIJS totaal
  Col T (19): Netto Inkoop
  Col U (20): Netto inkoop waarde
  Col V (21): Import (10,4%)
  Col W (22): Bruto Inkoop waarde
  Col X (23): verkoop prijs (incl.)
  Col Y (24): verkoop prijs (excl.)
  Col Z (25): verkoop waarde (excl.)
  Col AA (26): verkoop waarde (incl.)
  Col AB (27): VE-code
  Col AC (28): marge
"""
import openpyxl
import csv
import os
import sys
import re

OUTPUT_DIR = r'c:\Users\irvhu\Desktop\ventoz_import_csv'

BOX_NAMES = {'wit kleinste', 'wit middel', 'midden', 'groot', 'lang standaard',
             'lang groot', 'splash', 'mk ii', 'valk'}


def safe_str(val):
    if val is None:
        return ''
    s = str(val).strip().replace('\n', ' ').replace('\r', '')
    if s.lower() in ('none', 'null'):
        return ''
    if '#REF' in s:
        return ''
    return s


def safe_int(val, max_reasonable=99999):
    """Convert to int string. Values above max_reasonable are treated as errors (e.g. EAN in wrong column)."""
    if val is None:
        return ''
    try:
        n = int(float(str(val).replace(',', '.')))
        if abs(n) > max_reasonable:
            return ''
        return str(n)
    except (ValueError, TypeError):
        m = re.search(r'-?\d+', str(val))
        if m:
            n = int(m.group())
            if abs(n) > max_reasonable:
                return ''
            return str(n)
        return ''


def safe_float(val):
    if val is None:
        return ''
    try:
        s = str(val).replace(',', '.').strip()
        s = re.sub(r'[^\d.\-]', '', s)
        if not s:
            return ''
        return str(round(float(s), 2))
    except (ValueError, TypeError):
        return ''


def generate_voorraad_csv():
    """Parse the stock Excel with proper variant/product propagation.

    Two-pass approach:
      Pass 1: Build a lookup of artikelnummer -> product name from rows with explicit names.
      Pass 2: Process all rows, using the lookup to resolve orphan rows after separators.
    """
    path = r'c:\Users\irvhu\Desktop\2026 Ventoz VOORRAAD Zeilen.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    NAME_CORRECTIONS = {
        'blokart 2.1': 'blokart 2.0',
        'blokart 5,5': 'blokart 5.5',
    }

    # Pass 1: build artikelnummer -> product name lookup (most common name per art.nr)
    art_to_product = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        while len(cells) < 30:
            cells.append(None)
        product = safe_str(cells[1])
        artikelnr = safe_str(cells[3])
        if product and artikelnr and product.lower() not in ('totaal', 'totalen'):
            corrected = NAME_CORRECTIONS.get(product.lower(), product)
            key = artikelnr.strip()
            if key not in art_to_product:
                art_to_product[key] = corrected

    print(f'  Pass 1: {len(art_to_product)} artikelnummer->product mappings', file=sys.stderr)

    # Pass 2: process rows with proper propagation and orphan resolution
    rows_out = []
    current_category = ''
    current_product = ''
    prev_was_separator = False
    skipped = {'empty': 0, 'totaal': 0, 'boxes': 0, 'header': 0, 'no_data': 0, 'resolved': 0}

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row)
        while len(cells) < 30:
            cells.append(None)

        cat = safe_str(cells[0])
        product = safe_str(cells[1])
        kleur = safe_str(cells[2])
        artikelnr = safe_str(cells[3])
        ve_code = safe_str(cells[27])

        # Skip header-like rows
        if 'artikelnr' in product.lower() or 'voorraad' in product.lower():
            skipped['header'] += 1
            continue

        # Check for totally empty rows (separators between product groups)
        has_any_data = any(cells[i] is not None for i in [1, 2, 3, 7, 9, 13, 27])
        if not has_any_data:
            skipped['empty'] += 1
            prev_was_separator = True
            continue

        # Skip totaal rows (can appear in col B or col C)
        if kleur.lower() in ('totaal', 'totalen', 'total', 'subtotaal'):
            skipped['totaal'] += 1
            prev_was_separator = True
            continue
        if product.lower() in ('totaal', 'totalen', 'total', 'subtotaal'):
            skipped['totaal'] += 1
            prev_was_separator = True
            continue

        # Update category (propagates down)
        if cat:
            current_category = cat

        # Update product name (propagates down within same category group)
        if product:
            current_product = product
            prev_was_separator = False

        # Skip box/packaging rows
        effective_name = (product or kleur).lower()
        if effective_name in BOX_NAMES:
            skipped['boxes'] += 1
            continue
        if current_category.lower() == 'dozen':
            skipped['boxes'] += 1
            continue

        # Determine effective product name
        if product:
            effective_product = product
        elif prev_was_separator and artikelnr and artikelnr in art_to_product:
            # Orphan row after separator: resolve via artikelnummer lookup
            effective_product = art_to_product[artikelnr]
            skipped['resolved'] += 1
        else:
            effective_product = current_product

        # Apply known name corrections
        if effective_product.lower() in NAME_CORRECTIONS:
            effective_product = NAME_CORRECTIONS[effective_product.lower()]

        # Skip rows with truly no useful data
        if not effective_product and not artikelnr and not ve_code and not kleur:
            skipped['no_data'] += 1
            continue

        gewicht = safe_int(cells[4], max_reasonable=999999)
        gewicht_verp = safe_int(cells[5], max_reasonable=999999)
        ean = safe_str(cells[7]).replace('\n', '').strip()
        opmerking = safe_str(cells[8])
        voorraad = safe_int(cells[9])
        besteld = safe_int(cells[10])
        minimaal = safe_int(cells[12]) or safe_int(cells[11])
        inkoop = safe_float(cells[13])
        vervoer = safe_str(cells[14])
        vliegtuig_kosten = safe_float(cells[15])
        invoertax = safe_float(cells[16])
        inkoop_totaal = safe_float(cells[18])
        netto_inkoop = safe_float(cells[19])
        netto_inkoop_waarde = safe_float(cells[20])
        import_kosten = safe_float(cells[21])
        bruto_inkoop = safe_float(cells[22])
        verkoopprijs_incl = safe_float(cells[23])
        verkoopprijs_excl = safe_float(cells[24])
        verkoop_waarde_excl = safe_float(cells[25])
        verkoop_waarde_incl = safe_float(cells[26])
        marge = safe_float(cells[28])

        rows_out.append({
            'categorie': current_category,
            'product': effective_product,
            'kleur': kleur,
            'artikelnummer': artikelnr,
            'ean': ean,
            'code': ve_code,
            'voorraad': voorraad,
            'besteld': besteld,
            'minimaal': minimaal,
            'inkoop': inkoop,
            'vliegtuig_kosten': vliegtuig_kosten,
            'invoertax_admin': invoertax,
            'inkoop_totaal': inkoop_totaal,
            'netto_inkoop': netto_inkoop,
            'netto_inkoop_waarde': netto_inkoop_waarde,
            'import_kosten': import_kosten,
            'bruto_inkoop': bruto_inkoop,
            'verkoopprijs_incl': verkoopprijs_incl,
            'verkoopprijs_excl': verkoopprijs_excl,
            'verkoop_waarde_excl': verkoop_waarde_excl,
            'verkoop_waarde_incl': verkoop_waarde_incl,
            'vervoer': vervoer,
            'gewicht': gewicht,
            'gewicht_verpakking': gewicht_verp,
            'marge': marge,
            'opmerking': opmerking,
        })

    outpath = os.path.join(OUTPUT_DIR, 'ventoz_voorraad_zeilen.csv')
    fieldnames = ['categorie', 'product', 'kleur', 'artikelnummer', 'ean', 'code',
                  'voorraad', 'besteld', 'minimaal',
                  'inkoop', 'vliegtuig_kosten', 'invoertax_admin',
                  'inkoop_totaal', 'netto_inkoop', 'netto_inkoop_waarde',
                  'import_kosten', 'bruto_inkoop',
                  'verkoopprijs_incl', 'verkoopprijs_excl',
                  'verkoop_waarde_excl', 'verkoop_waarde_incl',
                  'vervoer', 'gewicht', 'gewicht_verpakking', 'marge', 'opmerking']
    with open(outpath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
        writer.writeheader()
        writer.writerows(rows_out)

    print(f'  Written {len(rows_out)} rows to {outpath}', file=sys.stderr)
    print(f'  Skipped: {skipped}', file=sys.stderr)

    # Print totals for verification
    total_stock = sum(int(r['voorraad'] or 0) for r in rows_out if r['voorraad'])
    total_rows_with_ve = sum(1 for r in rows_out if r['code'])
    products = set(r['product'].lower() for r in rows_out if r['product'])
    print(f'  Totaal voorraad: {total_stock} stk', file=sys.stderr)
    print(f'  Rijen met VE-code: {total_rows_with_ve}', file=sys.stderr)
    print(f'  Unieke producten: {len(products)}', file=sys.stderr)

    return rows_out


def generate_gewichten_csv():
    """Parse the weights Excel."""
    path = r'c:\Users\irvhu\Desktop\2026 Gewichten.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows_out = []
    current_category = ''

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        while len(cells) < 10:
            cells.append(None)

        cat = safe_str(cells[0])
        product = safe_str(cells[1])
        kleur = safe_str(cells[2])
        artikelnr = safe_str(cells[3])
        gewicht = safe_int(cells[4])
        gewicht_verp = safe_int(cells[5])
        ean = safe_str(cells[7]) if len(cells) > 7 else ''

        if cat:
            current_category = cat

        if not product and not artikelnr:
            continue
        if not gewicht and not gewicht_verp:
            continue

        rows_out.append({
            'categorie': current_category,
            'product': product,
            'kleur': kleur,
            'artikelnummer': artikelnr,
            'ean': ean,
            'gewicht': gewicht,
            'gewicht_verpakking': gewicht_verp,
        })

    outpath = os.path.join(OUTPUT_DIR, 'ventoz_gewichten.csv')
    fieldnames = ['categorie', 'product', 'kleur', 'artikelnummer', 'ean', 'gewicht', 'gewicht_verpakking']
    with open(outpath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
        writer.writeheader()
        writer.writerows(rows_out)

    print(f'  Written {len(rows_out)} rows to {outpath}', file=sys.stderr)


if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print('=== Generating CSV files for Ventoz import ===', file=sys.stderr)
    print(f'Output: {OUTPUT_DIR}\n', file=sys.stderr)

    print('1. Voorraad Zeilen...', file=sys.stderr)
    rows = generate_voorraad_csv()
    print(f'   Sample rows:', file=sys.stderr)
    for r in rows[:5]:
        print(f'   {r["product"]:30s} | {r["kleur"]:15s} | VR:{r["voorraad"]:>4s} | EAN:{r["ean"]:15s} | {r["code"]}', file=sys.stderr)
    print('', file=sys.stderr)

    print('2. Gewichten...', file=sys.stderr)
    generate_gewichten_csv()
    print('\nDone!', file=sys.stderr)
