#!/usr/bin/env python3
"""Generate SQL INSERT statements from Ventoz Excel files."""
import openpyxl
import sys

def esc(s):
    if s is None:
        return ''
    return str(s).strip().replace("'", "''")

def sql_str(s):
    v = esc(s)
    return f"'{v}'" if v else 'NULL'

def generate_ean_sql():
    wb = openpyxl.load_workbook(
        r'c:\Users\irvhu\Documents\Ventoz\VENTOZ\Documenten\EAN Codes\Alle 100 codes en produkten (2).xlsx',
        data_only=True
    )
    ws = wb['Blad1']
    
    lines = ['-- ============================================================']
    lines.append('-- EAN Registry - Alle EAN-codes met producttoewijzing')
    lines.append('-- ============================================================')
    lines.append('INSERT INTO ean_registry (artikelnummer, ean_code, product_naam, variant, kleur, opmerking, actief) VALUES')
    
    values = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        artnr = row[0]
        ean = row[1]
        if artnr is None or ean is None:
            continue
        try:
            artnr = int(float(str(artnr)))
        except (ValueError, TypeError):
            continue
        ean = str(ean).strip().replace('\n', '')
        if not ean.startswith('87'):
            continue
        prod = esc(row[2]) if len(row) > 2 else ''
        variant = esc(row[3]) if len(row) > 3 else ''
        kleur = esc(row[4]) if len(row) > 4 else ''
        opmerking = esc(row[8]) if len(row) > 8 else ''
        actief = 'false' if 'niet meer actief' in opmerking.lower() else 'true'
        opm_sql = f"'{opmerking}'" if opmerking else 'NULL'
        
        values.append(f"  ({artnr}, '{ean}', '{prod}', '{variant}', '{kleur}', {opm_sql}, {actief})")
    
    lines.append(',\n'.join(values))
    lines.append('ON CONFLICT (ean_code) DO UPDATE SET product_naam = EXCLUDED.product_naam, variant = EXCLUDED.variant, kleur = EXCLUDED.kleur, opmerking = EXCLUDED.opmerking;')
    lines.append('')
    return '\n'.join(lines)

def generate_sail_numbers_sql():
    wb = openpyxl.load_workbook(
        r'c:\Users\irvhu\Desktop\2026 Ventoz VOORRAAD Zeilnummers en Zeilletters.xlsx',
        data_only=True
    )
    ws = wb['Blad1']
    
    lines = ['-- ============================================================']
    lines.append('-- Zeilnummers en zeilletters - Initiële voorraad')
    lines.append('-- ============================================================')
    lines.append('INSERT INTO sail_numbers_letters (type, waarde, maat_mm, voorraad, opmerking) VALUES')
    
    values = []
    section_230 = True
    
    for row in ws.iter_rows(min_row=7, values_only=True):
        # Detect section switch
        val0 = str(row[0] or '').strip().lower()
        if '300 mm' in val0:
            section_230 = False
            continue
        if '230 mm' in val0:
            section_230 = True
            continue
        
        maat = 230 if section_230 else 300
        
        # Numbers (columns B, C)
        zeilnr = row[1]
        aantal_nr = row[2]
        if zeilnr is not None and str(zeilnr).strip():
            zeilnr_str = str(zeilnr).strip()
            try:
                if ',' in str(aantal_nr or ''):
                    stock = int(str(aantal_nr).split(',')[0].strip())
                elif isinstance(aantal_nr, (int, float)):
                    stock = int(aantal_nr)
                else:
                    stock_str = str(aantal_nr or '0').strip()
                    # Extract first number
                    import re
                    m = re.search(r'\d+', stock_str)
                    stock = int(m.group()) if m else 0
            except (ValueError, TypeError):
                stock = 0
            
            # Handle "6 en 9" as combined
            if 'en' in zeilnr_str.lower():
                waarde = zeilnr_str.replace(' ', '')
            else:
                waarde = zeilnr_str
            
            opmerking = esc(row[3]) if len(row) > 3 and row[3] else ''
            opm_sql = f"'{opmerking}'" if opmerking else 'NULL'
            values.append(f"  ('nummer', '{esc(waarde)}', {maat}, {stock}, {opm_sql})")
        
        # Letters (columns F, G)
        if len(row) > 6:
            zeilletter = row[5]
            aantal_lt = row[6]
            if zeilletter is not None and str(zeilletter).strip():
                letter_str = str(zeilletter).strip()
                if letter_str == '0' or letter_str.lower() in ('zeilletter', 'zeilletter ', ''):
                    continue
                try:
                    if isinstance(aantal_lt, (int, float)):
                        stock_lt = int(aantal_lt)
                    else:
                        stock_str = str(aantal_lt or '0').strip()
                        import re
                        m = re.search(r'\d+', stock_str)
                        stock_lt = int(m.group()) if m else 0
                except (ValueError, TypeError):
                    stock_lt = 0
                
                opmerking_lt = esc(row[7]) if len(row) > 7 and row[7] else ''
                opm_lt_sql = f"'{opmerking_lt}'" if opmerking_lt else 'NULL'
                values.append(f"  ('letter', '{esc(letter_str)}', {maat}, {stock_lt}, {opm_lt_sql})")
    
    lines.append(',\n'.join(values))
    lines.append('ON CONFLICT (type, waarde, maat_mm) DO UPDATE SET voorraad = EXCLUDED.voorraad;')
    lines.append('')
    return '\n'.join(lines)

if __name__ == '__main__':
    output = []
    output.append('-- ============================================================')
    output.append('-- Ventoz Sails - Initiële data voor voorraadbeheersysteem')
    output.append('-- Gegenereerd vanuit Excel bestanden')
    output.append('-- ============================================================')
    output.append('')
    
    print("Generating EAN registry...", file=sys.stderr)
    output.append(generate_ean_sql())
    
    print("Generating sail numbers/letters...", file=sys.stderr)
    output.append(generate_sail_numbers_sql())
    
    result = '\n'.join(output)
    
    outpath = r'c:\Users\irvhu\AndroidStudioProjects\ventoz-leads-app\supabase_initial_inventory_data.sql'
    with open(outpath, 'w', encoding='utf-8') as f:
        f.write(result)
    
    print(f"Written to {outpath}", file=sys.stderr)
    print(f"Total lines: {len(result.splitlines())}", file=sys.stderr)
