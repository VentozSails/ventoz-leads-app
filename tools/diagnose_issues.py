#!/usr/bin/env python3
"""Diagnose specific import problems by examining raw Excel data."""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\irvhu\Desktop\2026 Ventoz VOORRAAD Zeilen.xlsx', data_only=True)
ws = wb.active

# Print header for reference
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print("=== COLUMNS ===")
for i, h in enumerate(headers):
    if h is not None:
        letter = chr(65 + i) if i < 26 else chr(65 + (i//26-1)) + chr(65 + (i%26))
        print(f"  {letter:3s} ({i:2d}): {h}")

current_cat = ''
current_prod = ''

# Collect all rows with propagated names
print("\n\n" + "=" * 120)
print("ISSUE 1: BLOKART 5 / 5.5 / 5,5 area (rows around art=53,55,101)")
print("=" * 120)

for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)

    cat = str(cells[0] or '').strip()
    product = str(cells[1] or '').strip()
    kleur = str(cells[2] or '').strip()
    art = str(cells[3] or '').strip()

    if cat:
        current_cat = cat
    if product:
        current_prod = product

    effective = product if product else current_prod

    if 'blokart' in effective.lower() and ('5' in effective or art in ('53', '55', '101')):
        print(f"Row {ridx:3d}: cat='{current_cat:10s}' "
              f"prod_raw='{product:30s}' effective='{effective:30s}' "
              f"kleur='{kleur:25s}' art={art:5s} "
              f"VE={str(cells[27] or ''):10s} "
              f"voorraad={cells[9]} besteld={cells[10]} "
              f"ean={str(cells[7] or ''):20s}")

# Reset
current_cat = ''
current_prod = ''

print("\n\n" + "=" * 120)
print("ISSUE 2: HOBIE CAT 16 area (all rows)")
print("=" * 120)

for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)

    cat = str(cells[0] or '').strip()
    product = str(cells[1] or '').strip()

    if cat:
        current_cat = cat
    if product:
        current_prod = product

    effective = product if product else current_prod

    if 'hobie' in effective.lower() and '16' in effective:
        kleur = str(cells[2] or '').strip()
        art = str(cells[3] or '').strip()
        print(f"Row {ridx:3d}: prod_raw='{product:35s}' effective='{effective:35s}' "
              f"kleur='{kleur:25s}' art={art:5s} "
              f"VE={str(cells[27] or ''):10s} "
              f"ean='{str(cells[7] or ''):20s}' "
              f"voorraad={str(cells[9]):5s} besteld={str(cells[10]):5s} "
              f"inkoop={cells[13]}")

# Reset
current_cat = ''
current_prod = ''

print("\n\n" + "=" * 120)
print("ISSUE 3: LASER 4.7 / RADIAL 5.7 area")
print("=" * 120)

for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)

    cat = str(cells[0] or '').strip()
    product = str(cells[1] or '').strip()
    kleur = str(cells[2] or '').strip()
    art = str(cells[3] or '').strip()

    if cat:
        current_cat = cat
    if product:
        current_prod = product

    effective = product if product else current_prod

    if 'laser' in effective.lower() and ('4.7' in effective or '4,7' in effective or 
       'radial' in effective.lower() or '5.7' in effective or '5,7' in effective):
        print(f"Row {ridx:3d}: cat='{current_cat:10s}' "
              f"prod_raw='{product:30s}' effective='{effective:30s}' "
              f"kleur='{kleur:20s}' art={art:5s} "
              f"VE={str(cells[27] or ''):10s} "
              f"voorraad={cells[9]}")

# Also check: what product comes BEFORE and AFTER laser 4.7 in the sheet?
print("\n\n" + "=" * 120)
print("CONTEXT: Rows 55-90 (Laser area)")
print("=" * 120)
current_cat = ''
current_prod = ''

for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)

    cat = str(cells[0] or '').strip()
    product = str(cells[1] or '').strip()

    if cat:
        current_cat = cat
    if product:
        current_prod = product

    if 55 <= ridx <= 90:
        kleur = str(cells[2] or '').strip()
        art = str(cells[3] or '').strip()
        print(f"Row {ridx:3d}: cat='{current_cat:10s}' "
              f"prod_raw='{product:30s}' effective='{current_prod:30s}' "
              f"kleur='{kleur:20s}' art={art:5s} "
              f"VE={str(cells[27] or ''):10s} "
              f"voorraad={cells[9]}")
