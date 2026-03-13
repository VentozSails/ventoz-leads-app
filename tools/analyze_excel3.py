#!/usr/bin/env python3
"""Count how many product rows there are total, including variant rows without product name."""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\irvhu\Desktop\2026 Ventoz VOORRAAD Zeilen.xlsx', data_only=True)
ws = wb.active

total_data_rows = 0
rows_with_product = 0
rows_without_product = 0
rows_with_ve_code = 0
box_rows = 0
empty_rows = 0
totaal_rows = 0

box_names = ['wit kleinste', 'wit middel', 'midden', 'groot', 'lang standaard',
             'lang groot', 'splash', 'mk ii', 'valk']

current_product = ''
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)

    product = str(cells[1] or '').strip()
    kleur = str(cells[2] or '').strip()
    artikelnr = str(cells[3] or '').strip()
    ve_code = str(cells[27] or '').strip()
    voorraad = cells[9]

    # Check if it's totally empty
    has_data = any(cells[i] is not None for i in [1, 2, 3, 7, 9, 13, 27])
    if not has_data:
        empty_rows += 1
        continue

    # Check totaal
    if kleur.lower() == 'totaal' or product.lower() == 'totaal':
        totaal_rows += 1
        continue

    # Check boxes
    if product.lower() in box_names or (not product and kleur.lower() in box_names):
        box_rows += 1
        continue

    total_data_rows += 1
    if product:
        rows_with_product += 1
        current_product = product
    else:
        rows_without_product += 1

    if ve_code:
        rows_with_ve_code += 1

print(f'Total data rows (excl empty, totaal, boxes): {total_data_rows}')
print(f'  Rows with product name:    {rows_with_product}')
print(f'  Rows without product name: {rows_without_product}')
print(f'  Rows with VE-code:         {rows_with_ve_code}')
print(f'  Box rows (excluded):       {box_rows}')
print(f'  Totaal rows (excluded):    {totaal_rows}')
print(f'  Empty rows (excluded):     {empty_rows}')

# Now let's see how many UNIQUE VE codes there are
ve_codes = set()
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    ve = str(cells[27] or '').strip()
    kleur = str(cells[2] or '').strip()
    product = str(cells[1] or '').strip()
    if ve and kleur.lower() != 'totaal' and product.lower() not in box_names:
        ve_codes.add(ve)
print(f'\nUnique VE-codes: {len(ve_codes)}')

# How many rows have VE code + voorraad > 0?
with_stock = 0
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    ve = str(cells[27] or '').strip()
    kleur = str(cells[2] or '').strip()
    product = str(cells[1] or '').strip()
    voorraad = cells[9]
    if ve and kleur.lower() != 'totaal' and product.lower() not in box_names:
        if voorraad is not None and float(str(voorraad).replace(',','.') or '0') > 0:
            with_stock += 1
print(f'Rows with VE-code AND stock > 0: {with_stock}')

# And how many distinct items were imported: let's see how many had no VE code and some data
no_ve_with_data = 0
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    ve = str(cells[27] or '').strip()
    kleur = str(cells[2] or '').strip()
    product = str(cells[1] or '').strip()
    artikelnr = str(cells[3] or '').strip()
    has_data = any(cells[i] is not None for i in [1, 2, 3, 7, 9, 13])
    if not ve and has_data and kleur.lower() != 'totaal' and product.lower() not in box_names and kleur.lower() not in [x for x in box_names]:
        no_ve_with_data += 1
        if no_ve_with_data <= 10:
            print(f'  No VE, row {ridx}: product="{product}" kleur="{kleur}" art="{artikelnr}" voorraad={cells[9]}')

print(f'\nRows WITHOUT VE-code but with data: {no_ve_with_data}')
