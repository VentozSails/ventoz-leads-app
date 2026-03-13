#!/usr/bin/env python3
"""Verify exact column mapping by printing raw cell values for key rows."""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\irvhu\Desktop\2026 Ventoz VOORRAAD Zeilen.xlsx', data_only=True)
ws = wb.active

# Print header row
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print("=== HEADER ROW (Row 1) ===")
for i, h in enumerate(headers):
    if h is not None:
        letter = chr(65 + i) if i < 26 else chr(65 + (i//26-1)) + chr(65 + (i%26))
        print(f"  Col {letter} ({i:2d}): {h}")

# Hobie Cat 16 main rows (100-104 from previous analysis)
print("\n=== Hobie Cat 16 main - RAW cell values (rows 100-104) ===")
for ridx in range(100, 114):
    row = list(ws.iter_rows(min_row=ridx, max_row=ridx, values_only=True))[0]
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    
    # Only print rows that have data
    if cells[1] is not None or cells[2] is not None or cells[27] is not None:
        print(f"\nRow {ridx}:")
        for ci in [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,27,28]:
            letter = chr(65 + ci) if ci < 26 else chr(65 + (ci//26-1)) + chr(65 + (ci%26))
            val = cells[ci]
            if val is not None:
                print(f"  Col {letter} ({ci:2d}): {repr(val)}")

# Check: is the "besteld" column (K=10) actually containing EAN codes?
print("\n=== CHECKING: Does column K (10) contain EAN-like values? ===")
ean_in_besteld = 0
numeric_besteld = 0
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cells = list(row)
    while len(cells) < 30:
        cells.append(None)
    
    val_k = cells[10]
    val_h = cells[7]
    if val_k is not None:
        s = str(val_k)
        if len(s) > 10 and s.startswith('87'):
            ean_in_besteld += 1
            if ridx <= 115:
                print(f"  Row {ridx}: Col K (besteld)={val_k}, Col H (ean)={val_h}")
        else:
            numeric_besteld += 1

print(f"\nSummary: {ean_in_besteld} rows with EAN-like values in col K, {numeric_besteld} rows with normal numeric values")
